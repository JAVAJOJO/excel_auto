#!/usr/bin/env python3
"""
可配置 Excel 规则引擎 v2
支持：按列号或标题定位、搜索所有工作表
用法: python excel_rule_engine.py <规则文件.yml> <匹配值> [--dry-run]
"""

import argparse
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import yaml
import openpyxl
from openpyxl.utils import column_index_from_string

# ===================== 日志 & 备份配置 =====================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("rule_engine.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("RuleEngine")
BACKUP_DIR = Path("backup")

# ===================== 引擎核心 =====================
class ExcelRuleEngine:
    def __init__(self, rules: Dict[str, Any]):
        self.rules = rules
        self.task_name = rules.get("task_name", "Unnamed Task")
        self.file_groups: List[Dict] = rules.get("files", [])
        if not self.file_groups:
            raise ValueError("规则文件中必须包含 'files' 列表")

    def run(self, match_value: str, dry_run: bool = False):
        logger.info(f"{'='*50}")
        logger.info(f"任务: {self.task_name}  匹配值: {match_value}")
        if dry_run:
            logger.info("*** 预览模式，不会修改任何文件 ***")

        # 1. 非预览模式下先备份所有文件
        if not dry_run:
            self._backup_all()

        # 2. 逐个文件执行操作
        results = []
        for fg in self.file_groups:
            result = self._process_file(fg, match_value, dry_run)
            results.append(result)

        # 3. 汇总
        logger.info("="*50)
        logger.info("处理汇总:")
        for r in results:
            logger.info(f"  {r['file']}: 找到={r['found']}, 更新项={r['updates_applied']}")

        if dry_run:
            logger.info("预览完成，确认无误后去掉 --dry-run 参数执行。")

    def _backup_all(self):
        """备份所有规则中涉及的文件"""
        for fg in self.file_groups:
            path = Path(fg["file_path"])
            if path.exists():
                BACKUP_DIR.mkdir(exist_ok=True)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = BACKUP_DIR / f"{path.stem}_{ts}{path.suffix}"
                shutil.copy2(path, dest)
                logger.info(f"已备份: {path} -> {dest}")
            else:
                logger.warning(f"文件不存在，跳过备份: {path}")

    def _process_file(self, fg: Dict, match_value: str, dry_run: bool) -> Dict:
        file_path = Path(fg["file_path"])
        sheet_name = fg.get("sheet_name")          # "all" 表示所有工作表
        header_row = fg.get("header_row", 1)
        data_start = fg.get("data_start_row", header_row+1)
        operations = fg.get("operations", [])

        wb = openpyxl.load_workbook(file_path)
        
        # 确定要处理的工作表列表
        if sheet_name and sheet_name.lower() == "all":
            sheets = wb.sheetnames
        elif sheet_name:
            sheets = [sheet_name]
        else:
            sheets = [wb.active.title]

        total_found = False
        total_updates = 0

        for sheet in sheets:
            ws = wb[sheet]
            logger.info(f"\n处理工作表: [{ws.title}]")
            found, updates = self._process_sheet(ws, operations, match_value, header_row, data_start, dry_run)
            if found:
                total_found = True
            total_updates += updates

        if not dry_run and total_updates > 0:
            wb.save(file_path)
            logger.info("文件已保存")
        wb.close()

        return {"file": str(file_path), "found": total_found, "updates_applied": total_updates}

    def _process_sheet(self, ws, operations, match_value, header_row, data_start, dry_run) -> (bool, int):
        """处理单个工作表，返回 (是否找到匹配, 应用更新数)"""
        found_any = False
        updates_count = 0

        for op in operations:
            match_col_spec = op.get("match_column")
            if not match_col_spec:
                logger.warning("操作缺少 match_column，跳过")
                continue

            # 解析匹配列（支持列号或标题）
            match_col = self._resolve_column(ws, match_col_spec, header_row)
            if match_col is None:
                logger.error(f"无法解析匹配列: {match_col_spec}")
                continue

            # 查找匹配行
            matched_rows = []
            for row in range(data_start, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=match_col).value
                if cell_val and str(cell_val).strip() == match_value:
                    matched_rows.append(row)

            if not matched_rows:
                logger.warning(f"  在工作表 '{ws.title}' 的列 {match_col_spec} 中未找到 '{match_value}'")
                continue

            found_any = True
            for row in matched_rows:
                logger.info(f"  找到行 {row}，执行更新...")
                for upd in op.get("updates", []):
                    condition = upd.get("condition", "always")

                    # 解析目标列
                    target_spec = upd.get("column") or upd.get("col_index")
                    if not target_spec:
                        logger.error("更新项未指定列(column)或列号(col_index)")
                        continue
                    target_col = self._resolve_column(ws, target_spec, header_row)
                    if target_col is None:
                        logger.error(f"无法解析目标列: {target_spec}")
                        continue

                    current_val = ws.cell(row=row, column=target_col).value
                    do_update = False

                    if condition == "always":
                        do_update = True
                    elif condition == "is_empty":
                        if current_val is None or str(current_val).strip() == "":
                            do_update = True
                    elif condition == "equals":
                        compare_val = upd.get("compare_value")
                        if compare_val is not None and str(current_val).strip() == str(compare_val):
                            do_update = True
                    else:
                        logger.warning(f"未知条件: {condition}")

                    if do_update:
                        new_val = upd["new_value"]
                        if dry_run:
                            logger.info(f"    [预览] 列{target_spec}: '{current_val}' -> '{new_val}'")
                        else:
                            ws.cell(row=row, column=target_col).value = new_val
                            logger.info(f"    已更新列{target_spec}: '{current_val}' -> '{new_val}'")
                        updates_count += 1
                    else:
                        logger.info(f"    列{target_spec} 不满足条件 '{condition}'，跳过")
        return found_any, updates_count

    @staticmethod
    def _resolve_column(ws, spec: str, header_row: int) -> Optional[int]:
        """
        将列标识转换为列号（1-based）。
        - 如果 spec 是纯字母（如 "G", "AB"），直接转换为列号
        - 否则视为标题文本，在标题行查找
        """
        if spec.isalpha():          # 全字母，当作列字母处理
            try:
                return column_index_from_string(spec)
            except ValueError:
                pass
        # 按标题查找
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip() == spec.strip():
                return cell.column
        return None

# ===================== CLI =====================
def main():
    parser = argparse.ArgumentParser(description="可配置 Excel 规则引擎")
    parser.add_argument("rules_file", help="YAML 规则文件路径")
    parser.add_argument("match_value", help="要匹配的值（例如序列号）")
    parser.add_argument("--dry-run", action="store_true", help="预览模式，不修改文件")
    args = parser.parse_args()

    with open(args.rules_file, "r", encoding="utf-8") as f:
        rules = yaml.safe_load(f)

    engine = ExcelRuleEngine(rules)
    engine.run(args.match_value, args.dry_run)

if __name__ == "__main__":
    main()